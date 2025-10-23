# app.py
import streamlit as st
import pandas as pd
import numpy as np
import io
from pathlib import Path

# === Настройка внешнего вида ===
st.set_page_config(
    page_title="RadiaTool v1.9",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# CSS для максимального сходства с Tkinter-версией
st.markdown("""
<style>
    .main, .block-container {
        background-color: #e1dfdf !important;
        color: #444141 !important;
        font-family: "Segoe UI", sans-serif !important;
    }
    h1, h2, h3, h4, h5, h6 {
        color: #444141 !important;
        font-family: "Segoe UI", sans-serif !important;
    }
    .stButton > button {
        background-color: #955b67 !important;
        color: white !important;
        border: none !important;
        font-family: "Segoe UI", sans-serif !important;
        font-size: 14px !important;
        padding: 8px 16px !important;
        margin: 0 !important;
    }
    .stButton > button:hover {
        background-color: #263168 !important;
    }
    .stRadio > label, .stSelectbox > label {
        color: #444141 !important;
        font-family: "Segoe UI", sans-serif !important;
    }
    .stDataFrame {
        font-family: "Segoe UI", sans-serif !important;
    }
    div[data-testid="stHorizontalBlock"] > div {
        background-color: #dedede !important;
        padding: 0 !important;
        margin: 0 !important;
        border-radius: 0 !important;
    }
    .matrix-cell {
        text-align: center;
        background-color: #e6f3ff;
        border: 1px solid #ccc;
        padding: 0 !important;
        margin: 0 !important;
        height: 24px;
        width: 55px;
        font-family: "Segoe UI", sans-serif;
        font-size: 10px;
    }
    .matrix-cell-filled {
        background-color: #fff2cc !important;
    }
    .matrix-header {
        font-weight: bold;
        text-align: center;
        padding: 0 !important;
        margin: 0 !important;
        height: 24px;
        width: 55px;
        font-family: "Segoe UI", sans-serif;
        font-size: 12px;
    }
    .stNumberInput input {
        padding: 0 !important;
        margin: 0 !important;
        height: 24px;
        width: 55px;
        font-family: "Segoe UI", sans-serif;
        font-size: 10px;
    }
    .stFileUploader {
        padding: 0 !important;
        margin: 0 !important;
    }
    /* Убираем отступы между ячейками */
    .stColumns > div {
        padding: 0 !important;
        margin: 0 !important;
    }
    /* Стиль для верхних кнопок */
    .top-button {
        display: inline-block;
        background-color: #955b67 !important;
        color: white !important;
        border: none !important;
        font-family: "Segoe UI", sans-serif !important;
        font-size: 14px !important;
        padding: 8px 16px !important;
        margin: 0 5px !important;
        cursor: pointer;
    }
    .top-button:hover {
        background-color: #263168 !important;
    }
</style>
""", unsafe_allow_html=True)

# === Инициализация сессии ===
if "entry_values" not in st.session_state:
    st.session_state.entry_values = {}
if "connection" not in st.session_state:
    st.session_state.connection = "VK-правое"
if "radiator_type" not in st.session_state:
    st.session_state.radiator_type = "10"
if "bracket_type" not in st.session_state:
    st.session_state.bracket_type = "Настенные кронштейны"
if "radiator_discount" not in st.session_state:
    st.session_state.radiator_discount = 0.0
if "bracket_discount" not in st.session_state:
    st.session_state.bracket_discount = 0.0
if "show_tooltips" not in st.session_state:
    st.session_state.show_tooltips = False

# === Загрузка данных ===
@st.cache_data
def load_data():
    matrix_path = Path("data/Матрица.xlsx")
    brackets_path = Path("data/Кронштейны.xlsx")
    if not matrix_path.exists():
        st.error("❌ Файл 'Матрица.xlsx' не найден в папке data/")
        st.stop()
    if not brackets_path.exists():
        st.error("❌ Файл 'Кронштейны.xlsx' не найден в папке data/")
        st.stop()
    sheets = pd.read_excel(matrix_path, sheet_name=None, engine="openpyxl")
    brackets_df = pd.read_excel(brackets_path, engine="openpyxl")
    brackets_df['Артикул'] = brackets_df['Артикул'].astype(str).str.strip()
    for name, df in sheets.items():
        if name != "Кронштейны":
            df['Артикул'] = df['Артикул'].astype(str).str.strip()
            df['Вес, кг'] = pd.to_numeric(df['Вес, кг'], errors='coerce').fillna(0)
            df['Объем, м3'] = pd.to_numeric(df['Объем, м3'], errors='coerce').fillna(0)
    return sheets, brackets_df

sheets, brackets_df = load_data()

# === Вспомогательные функции ===
def parse_quantity(val):
    if not val:
        return 0
    try:
        if isinstance(val, (int, float)):
            return int(round(float(val)))
        val = str(val).strip()
        while val.startswith('+'): val = val[1:]
        while val.endswith('+'): val = val[:-1]
        if not val: return 0
        return sum(int(round(float(part.strip()))) for part in val.split('+') if part.strip())
    except:
        return 0

def calculate_brackets(radiator_type, length, height, bracket_type, qty=1):
    brackets = []
    if bracket_type == "Настенные кронштейны":
        if radiator_type in ["10", "11"]:
            brackets.extend([("К9.2L", 2*qty), ("К9.2R", 2*qty)])
            if 1700 <= length <= 2000:
                brackets.append(("К9.3-40", 1*qty))
        elif radiator_type in ["20", "21", "22", "30", "33"]:
            art_map = {300: "К15.4300", 400: "К15.4400", 500: "К15.4500", 600: "К15.4600", 900: "К15.4900"}
            if height in art_map:
                art = art_map[height]
                qty_br = 2*qty if 400 <= length <= 1600 else (3*qty if 1700 <= length <= 2000 else 0)
                if qty_br: brackets.append((art, qty_br))
    elif bracket_type == "Напольные кронштейны":
        if radiator_type in ["10", "11"]:
            art_map = {300: "КНС450", 400: "КНС450", 500: "КНС470", 600: "КНС470", 900: "КНС4100"}
            main_art = art_map.get(height)
            if main_art:
                brackets.append((main_art, 2*qty))
                if 1700 <= length <= 2000:
                    brackets.append(("КНС430", 1*qty))
        elif radiator_type == "21":
            art_map = {300: "КНС650", 400: "КНС650", 500: "КНС670", 600: "КНС670", 900: "КНС6100"}
            art = art_map.get(height)
            if art:
                if 400 <= length <= 1000: qty_br = 2*qty
                elif 1100 <= length <= 1600: qty_br = 3*qty
                elif 1700 <= length <= 2000: qty_br = 4*qty
                else: qty_br = 0
                if qty_br: brackets.append((art, qty_br))
        elif radiator_type in ["20", "22", "30", "33"]:
            art_map = {300: "КНС550", 400: "КНС550", 500: "КНС570", 600: "КНС570", 900: "КНС5100"}
            art = art_map.get(height)
            if art:
                if 400 <= length <= 1000: qty_br = 2*qty
                elif 1100 <= length <= 1600: qty_br = 3*qty
                elif 1700 <= length <= 2000: qty_br = 4*qty
                else: qty_br = 0
                if qty_br: brackets.append((art, qty_br))
    return brackets

def prepare_spec_data():
    spec_data = []
    bracket_temp = {}
    for (sheet_name, art), raw_val in st.session_state.entry_values.items():
        if not raw_val or sheet_name not in sheets:
            continue
        qty = parse_quantity(raw_val)
        if qty <= 0:
            continue
        df = sheets[sheet_name]
        product = df[df['Артикул'] == art]
        if product.empty:
            continue
        product = product.iloc[0]
        rad_type = sheet_name.split()[-1]
        price = float(product['Цена, руб'])
        disc = st.session_state.radiator_discount
        disc_price = round(price * (1 - disc / 100), 2)
        total = round(disc_price * qty, 2)
        name_parts = str(product['Наименование']).split('/')
        height = int(name_parts[-2].replace('мм', '').strip())
        length = int(name_parts[-1].replace('мм', '').strip().split()[0])
        conn_type = "VK" if "VK" in sheet_name else "K"
        spec_data.append({
            "№": len(spec_data) + 1,
            "Артикул": str(product['Артикул']),
            "Наименование": str(product['Наименование']),
            "Мощность, Вт": float(product.get('Мощность, Вт', 0)),
            "Цена, руб (с НДС)": price,
            "Скидка, %": disc,
            "Цена со скидкой, руб (с НДС)": disc_price,
            "Кол-во": qty,
            "Сумма, руб (с НДС)": total,
            "ConnectionType": conn_type,
            "RadiatorType": int(rad_type),
            "Height": height,
            "Length": length
        })
        if st.session_state.bracket_type != "Без кронштейнов":
            brackets = calculate_brackets(rad_type, length, height, st.session_state.bracket_type, qty)
            for art_b, qty_b in brackets:
                b_info = brackets_df[brackets_df['Артикул'] == art_b]
                if b_info.empty:
                    continue
                b_info = b_info.iloc[0]
                key = art_b.strip()
                if key not in bracket_temp:
                    bracket_temp[key] = {
                        "Артикул": art_b,
                        "Наименование": str(b_info['Наименование']),
                        "Цена, руб (с НДС)": float(b_info['Цена, руб']),
                        "Кол-во": 0,
                        "Сумма, руб (с НДС)": 0.0
                    }
                b_price = float(b_info['Цена, руб'])
                b_disc = st.session_state.bracket_discount
                b_disc_price = round(b_price * (1 - b_disc / 100), 2)
                bracket_temp[key]["Кол-во"] += qty_b
                bracket_temp[key]["Сумма, руб (с НДС)"] += round(b_disc_price * qty_b, 2)
    # Сортировка радиаторов
    spec_data.sort(key=lambda x: (0 if x["ConnectionType"] == "VK" else 1, x["RadiatorType"], x["Height"], x["Length"]))
    for i, item in enumerate(spec_data, 1):
        item["№"] = i
    # Добавление кронштейнов
    bracket_list = []
    for b in bracket_temp.values():
        b_disc = st.session_state.bracket_discount
        b_price = b["Цена, руб (с НДС)"]
        b_disc_price = round(b_price * (1 - b_disc / 100), 2)
        bracket_list.append({
            "№": len(spec_data) + len(bracket_list) + 1,
            "Артикул": b["Артикул"],
            "Наименование": b["Наименование"],
            "Мощность, Вт": 0.0,
            "Цена, руб (с НДС)": b_price,
            "Скидка, %": b_disc,
            "Цена со скидкой, руб (с НДС)": b_disc_price,
            "Кол-во": b["Кол-во"],
            "Сумма, руб (с НДС)": b["Сумма, руб (с НДС)"],
            "ConnectionType": "Bracket"
        })
    return pd.DataFrame(spec_data + bracket_list)

def save_excel_spec(df, correspondence_df=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Спецификация"

    headers = ["№", "Артикул", "Наименование", "Мощность, Вт", "Цена, руб (с НДС)", "Скидка, %", "Цена со скидкой, руб (с НДС)", "Кол-во", "Сумма, руб (с НДС)"]
    ws.append(headers)

    header_font = Font(name='Calibri', size=11, bold=True)
    data_font = Font(name='Calibri', size=11)
    alignment_center = Alignment(horizontal='center', vertical='center')
    alignment_left = Alignment(horizontal='left', vertical='center')
    thin = Side(border_style="thin")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = border

    for i, row in df.iterrows():
        power_val = "" if "Кронштейн" in str(row["Наименование"]) else row["Мощность, Вт"]
        ws.append([
            row["№"],
            str(row["Артикул"]),
            row["Наименование"],
            power_val,
            float(row["Цена, руб (с НДС)"]),
            float(row["Скидка, %"]),
            float(row["Цена со скидкой, руб (с НДС)"]),
            int(row["Кол-во"]),
            float(row["Сумма, руб (с НДС)"])
        ])
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=i+2, column=col)
            cell.font = data_font
            cell.border = border
            if col in [5,7,9]:
                cell.number_format = '#,##0.00'
                cell.alignment = alignment_center
            elif col == 4:
                cell.alignment = alignment_center
            elif col in [1,6,8]:
                cell.alignment = alignment_center
            else:
                cell.alignment = alignment_left

    total_row = len(df) + 2
    total_sum = df["Сумма, руб (с НДС)"].sum()
    rad_qty = df[df["Наименование"].str.contains("Радиатор", na=False)]["Кол-во"].sum()
    br_qty = df[df["Наименование"].str.contains("Кронштейн", na=False)]["Кол-во"].sum()
    ws.append(["Итого", "", "", "", "", "", "", f"{int(rad_qty)}/{int(br_qty)}", total_sum])
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.border = border
        cell.alignment = alignment_center
        if col in [5,7,9]:
            cell.number_format = '#,##0.00'

    # Вес и объем
    total_weight = 0.0
    total_volume = 0.0
    for _, row in df.iterrows():
        if "Кронштейн" in str(row["Наименование"]):
            continue
        art = str(row["Артикул"])
        qty = int(row["Кол-во"])
        for sheet_df in sheets.values():
            prod = sheet_df[sheet_df["Артикул"] == art]
            if not prod.empty:
                total_weight += float(prod.iloc[0]["Вес, кг"]) * qty
                total_volume += float(prod.iloc[0]["Объем, м3"]) * qty
                break
    ws.append([])
    ws.append([f"Суммарный вес радиаторов без учета упаковки и кронштейнов- {round(total_weight,1)} кг."])
    ws.merge_cells(start_row=total_row+2, start_column=1, end_row=total_row+2, end_column=9)
    ws.append([f"Суммарный объем радиаторов без учета упаковки и кронштейнов- {round(total_volume,3)} м3."])
    ws.merge_cells(start_row=total_row+3, start_column=1, end_row=total_row+3, end_column=9)

    col_widths = {'A':5,'B':12,'C':60,'D':15,'E':20,'F':10,'G':30,'H':10,'I':20}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    if correspondence_df is not None and not correspondence_df.empty:
        ws2 = wb.create_sheet("Таблица соответствия")
        ws2.append(list(correspondence_df.columns))
        for col in range(1, len(correspondence_df.columns)+1):
            cell = ws2.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = alignment_center
            cell.border = border
        for i, row in correspondence_df.iterrows():
            ws2.append(list(row))
            for col in range(1, len(correspondence_df.columns)+1):
                cell = ws2.cell(row=i+2, column=col)
                cell.font = data_font
                cell.border = border
                cell.alignment = alignment_center if col == 2 else alignment_left
        for col_idx, col_name in enumerate(correspondence_df.columns, 1):
            max_len = max(len(str(col_name)), correspondence_df[col_name].astype(str).map(len).max())
            ws2.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# === Интерфейс ===
st.title("RadiaTool v1.9")

# Верхнее меню (упрощённое)
col1, col2, col3 = st.columns([2, 3, 1])
with col1:
    if st.button("Создать спецификацию METEOR"):
        df = prepare_spec_data()
        if df.empty:
            st.warning("Нет данных для экспорта")
        else:
            excel_data = save_excel_spec(df)
            st.download_button("📥 Скачать Excel", excel_data, "Расчёт стоимости.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
with col2:
    upload_option = st.selectbox("", ["Выберите действие", "Загрузить спецификацию METEOR", "Загрузить CSV", "Загрузить иной спецификации"], index=0)
    if upload_option == "Загрузить спецификацию METEOR":
        uploaded_file = st.file_uploader("Загрузить спецификацию METEOR", type=["xlsx", "xls"], label_visibility="collapsed")
        if uploaded_file:
            st.success("Файл загружен (импорт в разработке)")
    elif upload_option == "Загрузить CSV":
        uploaded_file = st.file_uploader("Загрузить CSV", type=["csv"], label_visibility="collapsed")
        if uploaded_file:
            st.success("Файл загружен (импорт в разработке)")
    elif upload_option == "Загрузить иной спецификации":
        uploaded_file = st.file_uploader("Загрузить иной спецификации", type=["xlsx", "xls", "csv"], label_visibility="collapsed")
        if uploaded_file:
            st.success("Файл загружен (импорт в разработке)")
with col3:
    if st.button("Информация"):
        st.info("""
        **RadiaTool v1.9**  
        Программа для подбора радиаторов METEOR.  
        Поддержка: mt@laggartt.ru
        """)

# Основной контейнер
st.markdown("### Вид подключения")
conn_options = ["VK-правое", "VK-левое", "K-боковое"]
st.session_state.connection = st.radio("", conn_options, index=conn_options.index(st.session_state.connection), horizontal=True)

st.markdown("### Тип радиатора")
rad_types = ["10", "11", "30", "33"] if st.session_state.connection == "VK-левое" else ["10", "11", "20", "21", "22", "30", "33"]
st.session_state.radiator_type = st.radio("", rad_types, index=rad_types.index(st.session_state.radiator_type), horizontal=True)

# Матрица
st.markdown("#### длина радиаторов, мм")
sheet_name = f"{st.session_state.connection} {st.session_state.radiator_type}"
if sheet_name not in sheets:
    st.error(f"Лист '{sheet_name}' не найден")
else:
    df = sheets[sheet_name]
    lengths = list(range(400, 2100, 100))
    heights = [300, 400, 500, 600, 900]

    # Заголовки столбцов
    cols = st.columns(len(heights)+1)
    cols[0].markdown("<div class='matrix-header'>высота<br>радиаторов, мм</div>", unsafe_allow_html=True)
    for j, h in enumerate(heights):
        cols[j+1].markdown(f"<div class='matrix-header'>{h}</div>", unsafe_allow_html=True)

    has_any = any(st.session_state.entry_values.values())
    for i, l in enumerate(lengths):
        cols = st.columns(len(heights)+1)
        cols[0].markdown(f"<div class='matrix-header'>{l}</div>", unsafe_allow_html=True)
        for j, h in enumerate(heights):
            pattern = f"/{h}/{l}"
            match = df[df['Наименование'].str.contains(pattern, na=False)]
            if not match.empty:
                product = match.iloc[0]
                art = str(product['Артикул'])
                key = (sheet_name, art)
                current_val = st.session_state.entry_values.get(key, "")
                bg_class = "matrix-cell-filled" if current_val else ("matrix-cell" if has_any else "")
                with cols[j+1]:
                    new_val = st.text_input("", value=current_val, key=f"cell_{sheet_name}_{art}", label_visibility="collapsed")
                    st.session_state.entry_values[key] = new_val
                    if st.session_state.show_tooltips and new_val:
                        st.caption(f"Артикул: {art}")

# Нижняя панель
col1, col2, col3 = st.columns([2, 3, 2])
with col1:
    st.session_state.bracket_type = st.radio("Крепление", ["Настенные кронштейны", "Напольные кронштейны", "Без кронштейнов"], index=["Настенные кронштейны", "Напольные кронштейны", "Без кронштейнов"].index(st.session_state.bracket_type))
with col2:
    st.checkbox("Показывать параметры", value=st.session_state.show_tooltips, key="show_tooltips")
with col3:
    rad_disc = st.number_input("Скидка на радиаторы, %", min_value=0.0, max_value=100.0, value=st.session_state.radiator_discount, step=1.0, key="radiator_discount")
    br_disc = st.number_input("Скидка на кронштейны, %", min_value=0.0, max_value=100.0, value=st.session_state.bracket_discount, step=1.0, key="bracket_discount")

# Кнопки
col1, col2, col3 = st.columns([1, 4, 1])
with col1:
    if st.button("Предпросмотр"):
        df = prepare_spec_data()
        if not df.empty:
            st.dataframe(df, use_container_width=True)
        else:
            st.warning("Нет данных")
with col3:
    if st.button("Сброс"):
        st.session_state.entry_values = {}
        st.rerun()

# Спецификация под кнопкой Предпросмотр
st.markdown("### Спецификация")
df = prepare_spec_data()
if df.empty:
    st.info("Заполните матрицу, чтобы сгенерировать спецификацию.")
else:
    st.dataframe(df, use_container_width=True)
    total_sum = df["Сумма, руб (с НДС)"].sum()
    total_power = 0.0
    for _, row in df.iterrows():
        if "Кронштейн" not in str(row["Наименование"]):
            total_power += float(row["Мощность, Вт"]) * int(row["Кол-во"])
    st.markdown(f"**Суммарная мощность:** {total_power:.2f} Вт")
    st.markdown(f"**Сумма спецификации:** {total_sum:.2f} руб")